# 여러 플랏을 엑셀 하나의 시트에 특정 간격으로 집어 넣기

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# Create a sample DataFrame with random data
data = {
    f'x{i}': range(1, 11) for i in range(1, 17)
}
df = pd.DataFrame(data)

# Create a new Excel workbook
wb = Workbook()
ws = wb.active

# Create and save the plots as images and insert them into the worksheet
rows, cols = 4, 4
row_height = 20  # Adjust this value to increase/decrease the vertical spacing
col_width = 10   # Adjust this value to increase/decrease the horizontal spacing

for i in range(rows):
    for j in range(cols):
        fig, ax = plt.subplots()
        col_idx = i * cols + j + 1
        ax.plot(df[f'x{col_idx}'], df[f'x{col_idx}'] * 2)
        plt.title(f'Plot {col_idx}')
        plt.xlabel('x')
        plt.ylabel(f'y = 2 * x')
        
        # Save the plot as an image
        plt.savefig(f'plot_{col_idx}.png')
        plt.close()
        
        # Insert the image into the worksheet with adjusted offsets
        row_offset = i * (row_height + 5)  # Increase the row offset for larger vertical spacing
        col_offset = j * col_width  # Increase the column offset for larger horizontal spacing
        col_letter = get_column_letter(col_offset + 1)
        ws.add_image(Image(f'plot_{col_idx}.png'), f'{col_letter}{row_offset + 1}')

# Save the workbook
wb.save('plots.xlsx')
